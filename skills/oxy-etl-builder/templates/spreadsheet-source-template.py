"""
DLT Source Template (Spreadsheet)
=================================

Template for building DLT sources that extract data from spreadsheets.
Includes patterns for:
- File-based input
- Template auto-detection
- Entity mapping for multi-location files
- ETL metadata fields
"""

import dlt
import pendulum
from pathlib import Path
from typing import Iterator

import openpyxl

# Import template (adjust path)
# from .templates.income_statement import IncomeStatementTemplate


# =============================================================================
# Placeholder template - replace with actual import
# =============================================================================

class IncomeStatementTemplate:
    """Placeholder - replace with actual template import."""

    @classmethod
    def detect(cls, workbook) -> float:
        return 0.8

    @classmethod
    def get_entity_mapping(cls, workbook) -> dict[str, str]:
        return {sheet.title: sheet.title.lower() for sheet in workbook.worksheets}

    @classmethod
    def extract(cls, workbook, entity_mapping) -> Iterator[dict]:
        yield {"placeholder": True}


# =============================================================================
# Helper Functions
# =============================================================================

def load_workbook(file_path: str | Path):
    """
    Load Excel workbook with data values (not formulas).

    Args:
        file_path: Path to Excel file

    Returns:
        openpyxl Workbook object
    """
    return openpyxl.load_workbook(
        file_path,
        data_only=True,   # Get calculated values, not formulas
        read_only=True,   # Memory efficient for large files
    )


def _add_etl_metadata(record: dict, source_name: str) -> dict:
    """Add ETL provenance fields to a record."""
    return {
        **record,
        "_etl_source": source_name,
        "_etl_extracted_at": pendulum.now().isoformat(),
    }


# =============================================================================
# DLT Source
# =============================================================================

@dlt.source(name="income_statement")
def income_statement_source(
    file_path: str,
    entity_mapping: dict[str, str] | None = None,
    template_class: type | None = None,
) -> list:
    """
    DLT source for income statement spreadsheets.

    Args:
        file_path: Path to Excel file
        entity_mapping: Optional entity mapping (auto-detected if not provided)
        template_class: Optional template class (auto-detected if not provided)

    Returns:
        List of DLT resources
    """
    # Load workbook
    workbook = load_workbook(file_path)

    # Use provided template or default
    template = template_class or IncomeStatementTemplate

    # Auto-detect entity mapping if not provided
    if entity_mapping is None:
        entity_mapping = template.get_entity_mapping(workbook)

    # Return resources
    return [
        income_statement_facts(
            workbook=workbook,
            entity_mapping=entity_mapping,
            template=template,
            file_path=file_path,
        ),
    ]


# =============================================================================
# DLT Resources
# =============================================================================

@dlt.resource(
    name="income_statement_facts",
    write_disposition="merge",
    primary_key=["entity_id", "period", "category", "subcategory"],
)
def income_statement_facts(
    workbook,
    entity_mapping: dict[str, str],
    template: type,
    file_path: str,
) -> Iterator[dict]:
    """
    Extract income statement facts from spreadsheet.

    Args:
        workbook: openpyxl Workbook object
        entity_mapping: Map of sheet names to entity IDs
        template: Template class to use for extraction
        file_path: Original file path (for metadata)

    Yields:
        Fact records with ETL metadata
    """
    source_name = f"spreadsheet:{Path(file_path).name}"

    for record in template.extract(workbook, entity_mapping):
        yield _add_etl_metadata(record, source_name)


# =============================================================================
# Multi-Template Source
# =============================================================================

@dlt.source(name="spreadsheet_generic")
def spreadsheet_source(
    file_path: str,
    template_name: str | None = None,
) -> list:
    """
    Generic spreadsheet source with auto-detection.

    Args:
        file_path: Path to Excel file
        template_name: Optional template name (auto-detects if not provided)

    Returns:
        List of DLT resources
    """
    workbook = load_workbook(file_path)

    # Auto-detect template if not specified
    if template_name is None:
        # Import template registry (adjust path)
        # from .templates import detect_template
        # template_class, confidence = detect_template(workbook)

        # Placeholder
        template_class = IncomeStatementTemplate
        confidence = 0.8

        if template_class is None:
            raise ValueError(
                f"Could not auto-detect template for {file_path}. "
                "Please specify template_name explicitly."
            )

        print(f"Auto-detected template: {template_class.__name__} ({confidence:.0%})")
    else:
        # Look up template by name
        # template_class = get_template_by_name(template_name)
        template_class = IncomeStatementTemplate

    # Get entity mapping
    entity_mapping = template_class.get_entity_mapping(workbook)

    # Return appropriate resources
    return [
        spreadsheet_facts(
            workbook=workbook,
            entity_mapping=entity_mapping,
            template=template_class,
            file_path=file_path,
        ),
    ]


@dlt.resource(
    name="facts",
    write_disposition="merge",
    primary_key=["entity_id", "period", "category"],
)
def spreadsheet_facts(
    workbook,
    entity_mapping: dict[str, str],
    template: type,
    file_path: str,
) -> Iterator[dict]:
    """Generic fact extraction resource."""
    source_name = f"spreadsheet:{Path(file_path).name}"

    for record in template.extract(workbook, entity_mapping):
        yield _add_etl_metadata(record, source_name)
